import pandas as pd
import json
from copy import copy
import openpyxl
from openpyxl.styles import Alignment

def create_tc_excel(tc_data, output_filename):
    """Converts a list of test case data (Python dicts) into a formatted Excel file."""
    
    # 엑셀 파일 생성을 위한 ExcelWriter 객체 생성
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        # 각 테스트 케이스를 별도의 시트로 작성
        for i, tc in enumerate(tc_data):
            sheet_name = tc.get("id", f"TC_{i+1}")[:31] # 시트 이름은 31자 제한

            # 1. Pre-conditions, Notes를 상단에 기재
            # 줄바꿈을 유지하기 위해 셀 서식 설정
            preconditions_str = "\n".join(tc.get('preconditions', []))
            notes_str = tc.get('notes', '')
            header_data = {
                'Category': ['SRS ID', 'Title', 'Pre-conditions', 'Notes'],
                'Details': [tc.get('srs_id', ''), tc.get('title', ''), preconditions_str, notes_str]
            }
            header_df = pd.DataFrame(header_data)
            header_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=0)

            # 2. Test Steps를 표 형식으로 작성
            steps_df = pd.DataFrame(tc.get('test_steps', []))
            if not steps_df.empty:
                steps_df = steps_df.rename(columns={
                    'step': 'No.', 
                    'action': 'Action', 
                    'test_data': 'Test Data', 
                    'expected_result': 'Expected Result'
                })
                steps_df = steps_df[['No.', 'Action', 'Test Data', 'Expected Result']]
            steps_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(header_df) + 2)

            # 3. 보기 좋게 서식 조절
            worksheet = writer.sheets[sheet_name]
            # 컬럼 너비 자동 조절
            for column_cells in worksheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        # 셀 내용의 길이를 계산
                        cell_length = max(len(line) for line in str(cell.value).split('\n'))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column].width = min(adjusted_width, 100) # 최대 너비 100으로 제한
            
            # Details 컬럼의 줄바꿈 활성화
            for cell in worksheet['B']:
                # DeprecationWarning을 피하기 위해 copy 모듈 사용
                new_alignment = copy(cell.alignment)
                new_alignment.wrap_text = True
                cell.alignment = new_alignment

    print(f"Successfully created test case Excel file: {output_filename}")


def read_tc_excel(excel_file_path):
    """Reads test case data from a formatted Excel file and returns a list of TC dictionaries."""
    all_tcs = []
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        print(f"❌ Error: Excel file not found at {excel_file_path}")
        return []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        tc_id = sheet_name # Sheet name is the TC ID

        tc_data = {
            "id": tc_id,
            "srs_id": "",
            "title": "",
            "preconditions": [],
            "test_steps": [],
            "notes": ""
        }

        # Read SRS ID, Title, Pre-conditions and Notes (from rows 1, 2, 3, 4, column B)
        srs_id_val = worksheet['B1'].value
        if srs_id_val:
            tc_data["srs_id"] = str(srs_id_val).strip()

        title_val = worksheet['B2'].value
        if title_val:
            tc_data["title"] = str(title_val).strip()

        preconditions_str = worksheet['B3'].value
        if preconditions_str:
            tc_data["preconditions"] = [line.strip() for line in preconditions_str.split('\n') if line.strip()]
        
        notes_str = worksheet['B4'].value # Now in B3
        if notes_str:
            tc_data["notes"] = notes_str

        # Read Test Steps (starting from row 6, assuming header is in row 5)
        # Find the header row for steps dynamically
        steps_header_row = -1
        # Start searching for 'No.' from row 1, but expect it to be around row 5 or 6
        for r_idx in range(1, worksheet.max_row + 1):
            if worksheet.cell(row=r_idx, column=1).value == 'No.':
                steps_header_row = r_idx
                break
        
        if steps_header_row != -1:
            col_map = {}
            for c_idx in range(1, worksheet.max_column + 1):
                header_val = worksheet.cell(row=steps_header_row, column=c_idx).value
                if header_val == 'No.': col_map['step'] = c_idx
                elif header_val == 'Action': col_map['action'] = c_idx
                elif header_val == 'Test Data': col_map['test_data'] = c_idx
                elif header_val == 'Expected Result': col_map['expected_result'] = c_idx
            
            for r_idx in range(steps_header_row + 1, worksheet.max_row + 1):
                step_num = worksheet.cell(row=r_idx, column=col_map.get('step')).value
                if step_num is None: # Stop if no step number
                    continue

                action = worksheet.cell(row=r_idx, column=col_map.get('action')).value
                test_data = worksheet.cell(row=r_idx, column=col_map.get('test_data')).value
                expected_result = worksheet.cell(row=r_idx, column=col_map.get('expected_result')).value

                tc_data["test_steps"].append({
                    "step": step_num,
                    "action": action if action is not None else "",
                    "test_data": test_data if test_data is not None else "",
                    "expected_result": expected_result if expected_result is not None else ""
                })
        
        all_tcs.append(tc_data)

    return all_tcs


if __name__ == '__main__':
    # LLM이 생성할 것으로 예상되는 가상의 데이터 (Python 딕셔너리 형태)
    mock_data = [
        {
            "id": "TC-JIRA-001",
            "title": "Jira 연동을 통한 테스트 계획 생성 성공 시나리오",
            "preconditions": [
                "백엔드 서비스에 유효한 Jira 연동 정보(서버 URL, 사용자 이메일, API 토큰)가 하드코딩 또는 환경 변수로 설정되어 있다.",
                "사용자가 애플리케이션에 접근했다."
            ],
            "test_steps": [
                {
                    "step": 1,
                    "action": "'새 계획 생성' 페이지에서 제품명과 버전을 입력하고 '테스트 케이스 불러오기' 버튼을 클릭한다.",
                    "test_data": "(유효한 제품명/버전)",
                    "expected_result": "백엔드 서비스가 Jira API와 정상적으로 통신하여 테스트 케이스 목록을 조회해온다.\n오류 없이 테스트 케이스 목록이 화면에 표시된다"
                },
                {
                    "step": 2,
                    "action": "1개 이상의 테스트 케이스를 선택하고 '테스트 계획 생성' 및 Dry Run 모달의 '생성' 버튼을 클릭한다.",
                    "test_data": "-",
                    "expected_result": "백엔드 서비스가 Jira API와 정상적으로 통신하여 Epic 및 Test Run 이슈들을 생성한다.\n'생성 완료' 페이지로 이동한다."
                }
            ],
            "notes": "만약 인증 정보가 잘못된 경우, \"Jira 인증에 실패했습니다\" 와 같은 명확한 에러 메시지가 사용자에게 표시되어야 한다."
        }
    ]

    output_file = "generated_test_cases.xlsx"
    # 이제 함수는 JSON 문자열이 아닌 파이썬 딕셔너리 리스트를 직접 받음
    create_tc_excel(mock_data, output_file)